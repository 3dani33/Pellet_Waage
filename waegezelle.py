from openpyxl import load_workbook, Workbook
import datetime
import time

FILENAME = 'calibration_data.xlsx'
SHEETNAME = 'data'

try:
	workbook = load_workbook(filename=FILENAME)
	print('loaded {}.'.format(FILENAME))
except FileNotFoundError:
	workbook = Workbook()
	print('{} not found, will be created after the program is finished.'.format(FILENAME))

sheet = workbook.active  # gets first sheet
sheet.title = SHEETNAME  # set name
print('writing to sheet "{}"'.format(sheet.title))

sheet['B1'] = 'hey'

column = 1
while sheet.cell(row=1, column=column).value is not None:
	print('column {}: {}'.format(column, sheet.cell(row=1, column=column).value))
	column += 1
print('writing to first empyty column: {}'.format(column))

# save date and time
sheet.cell(row=1, column=column, value=datetime.datetime.now().replace(microsecond=0))


# import hx711
print('setting up the hx711 and GPIO pins.')
import RPi.GPIO as GPIO
from hx711 import HX711
hx = HX711(5, 6)
hx.set_reading_format("MSB", "MSB")
hx.set_reference_unit(1)
hx.reset()
# TO TARE OR NOT TO TARE, THAT IS THE QUESTION…
# print('taring the hx711.')
# hx.tare()
# print("Tare done! Add weight now...")

row = 2

while True:
	try:
		data = hx.get_weight(5)  # averaging over 5 values

		sheet.cell(row=row, column=column, value=data)
		row += 1
		print('data: {}, row: {}'.format(data, row))

		# save every time?
		# -> workbook.save(filename=FILENAME)

		hx.power_down()  # needed?
		hx.power_up()
		time.sleep(0.1)
	except (KeyboardInterrupt, SystemExit):
		print('\nstopping the program.')
		break

workbook.save(filename=FILENAME)
print('saved the file as {}.'.format(FILENAME))
GPIO.cleanup()
print('exiting.')